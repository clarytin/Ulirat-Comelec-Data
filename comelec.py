from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, ElementClickInterceptedException
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import time
import traceback


REGION = 2
PROVINCE = 3
MUNICIPALITY = 4

PRESIDENT = '+5587'
VICE_PRES = '+5588'
SENATOR = '+5589'
PARTY_LIST = '+11172'
MEMBER_HOR = 'MEMBER, HOUSE OF REPRESENTATIVES'
GOVERNOR = 'PROVINCIAL GOVERNOR'
VICE_GOV = 'PROVINCIAL VICE-GOVERNOR'
SANG_PANLA = 'PANLALAWIGAN'
MAYOR = 'MAYOR'
VICE_MAYOR = 'VICE-MAYOR'
SANG_BAYAN = 'SANGGUNIANG BAYAN'
SANG_PANLU = 'PANLUNGSOD'

positions = [PRESIDENT, VICE_PRES, SENATOR, PARTY_LIST, 
             MEMBER_HOR, GOVERNOR, VICE_GOV, SANG_PANLA, MAYOR, VICE_MAYOR]

title = {PRESIDENT: "PRESIDENT PHILIPPINES",
         VICE_PRES: "VICE PRESIDENT PHILIPPINES",
         SENATOR: "SENATOR PHILIPPINES",
         PARTY_LIST: "PARTY LIST PHILIPPINES",
         MEMBER_HOR: "MEMBER, HOUSE OF REPRESENTATIVES",
         GOVERNOR: "PROVINCIAL GOVERNOR",
         VICE_GOV: "PROVINCIAL VICE-GOVERNOR",
         SANG_PANLA: "MEMBER, SANGGUNIANG PANLALAWIGAN",
         MAYOR: "MAYOR",
         VICE_MAYOR: "VICE-MAYOR",
         SANG_BAYAN: "MEMBER, SANGGUNIANG BAYAN",
         SANG_PANLU: "MEMBER, SANGGUNIANG PANLUNGSOD"}

# Number 4 is skipped on the website
region_idx = [2, 3, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]
global driver
global soup
global wb


def setup(): 
    chrome_options = Options()
    chrome_options.add_experimental_option("detach", True)
    cService = webdriver.ChromeService(executable_path='/Users/clarissa/Developer/Projects/Ulirat-Comelec-Data/chromedriver')
   
    driver = webdriver.Chrome(service = cService, options=chrome_options)
    driver.get('https://2022electionresults.comelec.gov.ph/#/coc/0')
    driver.implicitly_wait(5)

    return driver

def choose_area(reg, prov, mun):
    click_filter_btn(REGION)
    click_dropdown(REGION, reg)
    click_filter_btn(PROVINCE)
    click_dropdown(PROVINCE, prov)
    click_filter_btn(MUNICIPALITY)
    click_dropdown(MUNICIPALITY, mun)


def click_filter_btn(area):
    xpath = '//*[@id="container"]/ui-view/div/div/div[1]/nav/div/ul/li/div[4]/div'
    xpath += '[' + str(area) + ']'
    xpath += '/nav-filter/div/span/div/div/span'
    region_filter_button = driver.find_element(By.XPATH, xpath)
    region_filter_button.click()

def click_dropdown(area, option): 
    xpath = '//*[@id="container"]/ui-view/div/div/div[1]/nav/div/ul/li/div[4]/div'
    xpath += '[' + str(area) + ']'
    xpath += '/nav-filter/div/span/div/div/div[2]/ul/li'
    xpath += '[' + str(option) + ']'
    elem = driver.find_element(By.XPATH, xpath)
    driver.execute_script("arguments[0].click();", elem)

def clean(data):
    return data.getText()[1:-1]


def get_data_div(position):
    if position in [PRESIDENT, VICE_PRES, SENATOR, PARTY_LIST]:
        div_id = "'resultDiv.'" + position
        results_div = soup.find(id=div_id)
        return results_div
    
    try:
        div = soup.select(':-soup-contains("' + position + '")')[-1]
    except IndexError:
        if position == SANG_BAYAN:
            # someone misspelled this
            div = soup.select(':-soup-contains("SANGUNIANG BAYAN")')[-1]
 
    # Previous line goes to VICE-MAYOR if position is MAYOR
    # Manually navigate to the div I need
    if position == MAYOR:
        div = div.parent.parent
        div = div.parent.parent.parent.find_previous_sibling()
        div = div.findChild("span", {'class': 'ng-binding'})
    
    return div.parent.parent.find_next_sibling()

def get_vote_data(position):
    results_div = get_data_div(position).findChild("div")
    results = results_div.findChildren("div", {'class': 'candidate-result'})

    df = pd.DataFrame(columns=['Candidate','Votes','Percentage'])

    for i in range(0, len(results), 3):
        row = pd.DataFrame({'Candidate':[results[i].text],
                            'Votes': [clean(results[i+1])],
                            'Percentage': [clean(results[i+2])]})
        df = pd.concat([df, row], ignore_index=True)

    return df

def get_stats(position):
    stats_div = get_data_div(position).findChild("div", {'id': 'generalStatisticsVoters'})
    stats = stats_div.findChildren("div", {'class': 'ng-binding'})

    df = pd.DataFrame({stats[1].text: [stats[2].text],
                       stats[3].text: [stats[4].text],
                       stats[5].text: [stats[6].text],
                       stats[7].text: [stats[8].text]})

    return df

def get_name(area):
    xpath = ' //*[@id="container"]/ui-view/div/div/div[1]/nav/div/ul/li/div[4]/div'
    xpath += '[' + str(area) + ']'
    xpath += '/nav-filter/div/span/span'
    name = driver.find_element(By.XPATH, xpath)
    return name.text

def init_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DELETE"
    wb.save(filename)
    return wb

def write_title(filename, position, curr_row):
    wb = openpyxl.load_workbook(filename)
    ws = wb[get_name(MUNICIPALITY)]
    ws.cell(column=1, row=curr_row, value=title[position])
    wb.save(filename)

def write_data(filename, target_region):
    curr_row = 1
    curr_positions = [i for i in positions]
    if "CITY" in target_region:
        curr_positions.append(SANG_PANLU)
    else:
        curr_positions.append(SANG_BAYAN)

    for position in curr_positions:
        votes = get_vote_data(position)
        stats = get_stats(position)

        write_title(filename, position, curr_row)
        curr_row += 1

        with pd.ExcelWriter(filename, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            votes.to_excel(writer, sheet_name=get_name(MUNICIPALITY), startrow=curr_row, index=False)
            curr_row += votes.shape[0] + 2
            stats.to_excel(writer, sheet_name=get_name(MUNICIPALITY), startrow=curr_row, index=False) 
            curr_row += stats.shape[0] + 4        

def get_region():
    return driver.find_element(By.XPATH, '//*[@id="container"]/ui-view/div/div/div[1]/nav/div/ul/li/div[4]/div[4]/nav-filter/div/span/span').text

def create_worksheet(muni):
    if muni == 1:
        wb = init_workbook()
        wb.create_sheet(get_name(MUNICIPALITY)) 
        del wb["DELETE"]
    else:
        wb = openpyxl.load_workbook(filename)
        wb.create_sheet(get_name(MUNICIPALITY)) 

    wb.save(filename)


        
failures = 0

for reg in region_idx:
    prov = 1

    while(True):
        muni = 1
        failures = 0
        
        while(True):
            if failures == 5:
                    break
            try:
                driver = setup()
                choose_area(reg, prov, muni)
                time.sleep(3)

                soup = BeautifulSoup(driver.page_source, 'html.parser')
                filename = 'data/' + get_name(REGION) + '/' + get_name(PROVINCE) + '.xlsx'

                create_worksheet(muni)
                write_data(filename, get_region())

                muni += 1
                failures = 0

            except NoSuchElementException:
                print("may be end")
                failures += 1
                continue

            except Exception as e: 
                print(e)        
                #print(traceback.format_exc())
                print("Failure " + str(failures) + " on region: " + \
                    str(reg) + ", province: " + str(prov) + ", muni: " + str(muni))
                
                if muni != 1 and failures != 5:
                    wb = openpyxl.load_workbook(filename)
                    del wb[get_name(MUNICIPALITY)]
                    wb.save(filename) 

                failures += 1
                continue

            finally:
                driver.close()
                

        if muni == 1:
            break

        prov += 1
